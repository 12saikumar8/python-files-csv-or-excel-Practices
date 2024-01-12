# from openpyxl import Workbook

# workbook = Workbook()
# sheet=workbook.active

# sheet["A1"]="Good"
# sheet["B1"]="Luck"

# workbook.save(filename="hello_world.xlsx")

from openpyxl import load_workbook
sai=load_workbook(filename="reviews-sample.xlsx")
# ws2 = sai.create_sheet(title = 'California')
# print(sai.sheetnames)
sheet=sai.active
print(sheet.title)




